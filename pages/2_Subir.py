"""
Página Subir - Clean Architecture
Simple upload and analysis page using only src/
"""

import streamlit as st
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# POLISH-002 FIX: Import constants for consistent configuration
try:
    from src.infrastructure.external_services.ai_engine_constants import AIEngineConstants
    CONSTANTS_AVAILABLE = True
except ImportError:
    CONSTANTS_AVAILABLE = False

# Add src to path
current_dir = Path(__file__).parent.parent
if str(current_dir) not in sys.path:
    sys.path.insert(0, str(current_dir))

# Import Clean Architecture components + CSS utilities
try:
    from src.shared.exceptions.archivo_exception import ArchivoException
    from src.shared.exceptions.ia_exception import IAException
    
    # HIGH-003 FIX: Centralized CSS loading strategy (no more import redundancy)
    try:
        # Primary: Enhanced CSS loader with all features
        from src.presentation.streamlit.enhanced_css_loader import (
            ensure_css_loaded, inject_page_css
        )
        
        # Try to import utility functions from basic loader
        try:
            from src.presentation.streamlit.css_loader import glass_card, metric_card
            CSS_UTILS_AVAILABLE = True
        except ImportError:
            # HIGH-003 FIX: Create safe fallback implementations
            def glass_card(content: str) -> None:
                """Fallback glass card implementation"""
                st.markdown(
                    f'<div class="glass-card" style="'
                    f'background: rgba(255, 255, 255, 0.08); '
                    f'backdrop-filter: blur(16px); '
                    f'border-radius: 16px; '
                    f'padding: 1rem; '
                    f'border: 1px solid rgba(255, 255, 255, 0.15);">'
                    f'{content}</div>', 
                    unsafe_allow_html=True
                )
            
            def metric_card(title: str, value: str) -> None:
                """Fallback metric card implementation"""
                st.metric(title, value)
            
            CSS_UTILS_AVAILABLE = True
            logger.info("⚠️ Using fallback CSS utility functions")
        
        # Load CSS with enhanced loader
        try:
            ensure_css_loaded()
            inject_page_css('upload')
            inject_page_css('analysis')
            ENHANCED_CSS_LOADED = True
        except Exception as css_error:
            logger.error(f"❌ Enhanced CSS loading failed: {str(css_error)}")
            ENHANCED_CSS_LOADED = False
        
    except ImportError as e:
        # HIGH-003 FIX: Complete fallback to basic CSS system
        logger.warning("⚠️ Enhanced CSS not available, using basic fallback")
        try:
            from src.presentation.streamlit.css_loader import load_css, glass_card, metric_card
            load_css()
            CSS_UTILS_AVAILABLE = True
            ENHANCED_CSS_LOADED = False
        except ImportError:
            logger.error("❌ No CSS system available")
            # Create minimal fallbacks
            def glass_card(content: str) -> None:
                st.markdown(content, unsafe_allow_html=True)
            def metric_card(title: str, value: str) -> None:
                st.metric(title, value)
            CSS_UTILS_AVAILABLE = False
            ENHANCED_CSS_LOADED = False
        
except ImportError as e:
    st.error(f"Error importando Clean Architecture: {str(e)}")
    CSS_UTILS_AVAILABLE = False


@st.fragment(run_every=0.5)
def live_batch_progress():
    """
    OPTIMIZATION: Real-time progress updates using Streamlit fragments
    Updates every 0.5s without full page rerun (Streamlit-native pattern)
    """
    if 'batch_progress_data' not in st.session_state:
        return
    
    progress_data = st.session_state.batch_progress_data
    action = progress_data.get('action', 'unknown')
    
    # Show progress container
    with st.container():
        st.markdown("### 🚀 Progreso del Análisis IA")
        
        if action == 'start':
            total_batches = progress_data.get('total_batches', 0)
            total_comments = progress_data.get('total_comments', 0)
            
            if total_comments > 0:  # Only show if we have real data
                st.progress(0.0, text=f"🚀 Iniciando análisis: {total_comments} comentarios en {total_batches} lotes")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("📊 Comentarios", total_comments)
                with col2:
                    st.metric("📦 Lotes", total_batches)
                with col3:
                    st.metric("⚡ Modo", "AsyncIO" if total_batches > 2 else "Secuencial")
            else:
                st.info("🔄 Preparando análisis...")
                
        elif action in ['batch_start', 'batch_success']:
            current_batch = progress_data.get('current_batch', 0)
            total_batches = progress_data.get('total_batches', 1)
            progress_pct = progress_data.get('progress_percentage', 0.0)
            confidence = progress_data.get('confidence', 0.0)
            
            # Progress bar with dynamic text
            status_icon = "✅" if action == 'batch_success' else "🔄"
            status_text = "Completado" if action == 'batch_success' else "Procesando"
            st.progress(progress_pct / 100, text=f"{status_icon} {status_text}: Lote {current_batch}/{total_batches}")
            
            # Real-time metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📈 Progreso", f"{progress_pct:.1f}%")
            with col2:
                if confidence > 0:
                    st.metric("🎯 Confianza", f"{confidence:.2f}")
                else:
                    st.metric("🎯 Confianza", "---")
            with col3:
                processing_mode = "AsyncIO" if total_batches > 2 else "Secuencial"
                st.metric("⚡ Modo", processing_mode)
                
        elif action == 'batch_failure':
            current_batch = progress_data.get('current_batch', 0)
            total_batches = progress_data.get('total_batches', 1)
            progress_pct = progress_data.get('progress_percentage', 0.0)
            reason = progress_data.get('reason', 'Error desconocido')
            
            st.progress(progress_pct / 100, text=f"❌ Error en lote {current_batch}/{total_batches}")
            st.error(f"**Lote {current_batch}**: {reason}")
            
        else:
            # Default state
            st.info("🤖 Análisis en preparación...")
            st.progress(0.0, text="Inicializando sistema de IA...")

def _run_analysis(uploaded_file, analysis_type):
    """Run pure IA analysis using maestro system with enhanced fragment-based progress tracking"""
    
    # OPTIMIZATION: Fragment-based progress display (no containers needed)
    # The live_batch_progress fragment will handle all display logic
    st.markdown("### 🚀 Análisis en Progreso")
    
    # Initialize session state for fragment communication
    if 'batch_progress_data' not in st.session_state:
        st.session_state.batch_progress_data = {'action': 'initialize'}
    
    # Display live progress using fragment (auto-updates every 0.5s)
    live_batch_progress()
    
    # Create enhanced progress callback for fragment communication
    def create_enhanced_progress_callback():
        """Create callback that updates session state for fragment display"""
        def update_progress(progress_data):
            # Store in session state for fragment access
            st.session_state.batch_progress_data = progress_data.copy()
            
            # Add processing mode info
            if 'total_batches' in progress_data:
                total_batches = progress_data['total_batches']
                st.session_state.batch_progress_data['processing_mode'] = (
                    'AsyncIO' if total_batches > 2 else 'Secuencial'
                )
            
            # Fragment will auto-update based on session state
            # No manual container updates needed
        
        return update_progress
    
    # Create the enhanced progress callback
    progress_callback = create_enhanced_progress_callback()
    
    # Initialize progress state (fragment will display)
    st.session_state.batch_progress_data = {
        'action': 'start',
        'total_batches': 0,
        'total_comments': 0,
        'current_batch': 0,
        'progress_percentage': 0.0
    }
    
    # Mark analysis as in progress
    st.session_state.ai_analysis_in_progress = True
        
    try:
        # Import session validator for robust checking
        from src.presentation.streamlit.session_validator import is_ia_system_ready
        
        # Pure IA analysis - validate system is ready
        if not is_ia_system_ready():
            st.error("Sistema IA no está disponible. Verifica configuración de OpenAI API key.")
            return
        
        # Get caso uso maestro with progress callback
        # Verificar disponibilidad del contenedor de dependencias
        if not ('contenedor' in st.session_state and st.session_state.contenedor):
            st.error("❌ Contenedor de dependencias no está disponible")
            st.info("💡 La aplicación no se inicializó correctamente")
            st.info("🔄 Recargar la página o verificar configuración")
            return
        
        # Obtener caso de uso maestro WITH progress callback
        try:
            caso_uso_maestro = st.session_state.contenedor.obtener_caso_uso_maestro(progress_callback)
        except Exception as e:
            st.error(f"❌ Error obteniendo sistema de análisis: {str(e)}")
            st.info("💡 Posibles causas:")
            st.info("- OpenAI API key no configurada o inválida")
            st.info("- Problemas de conexión a internet") 
            st.info("- Configuración incorrecta en variables de entorno")
            return
            
        if not caso_uso_maestro:
            st.error("❌ Sistema de análisis IA no está disponible")
            st.info("🔧 Configurar variables de entorno requeridas:")
            st.code("""
OPENAI_API_KEY=your-api-key-here
OPENAI_MODEL=gpt-4o-mini  
OPENAI_MAX_TOKENS=8000
MAX_COMMENTS_PER_BATCH=20
            """)
            return
        
        # Verificar disponibilidad del analizador IA específicamente
        try:
            analizador = st.session_state.contenedor.obtener_analizador_maestro_ia()
            if not analizador.es_disponible():
                st.warning("⚠️ Sistema IA en modo degradado - API no disponible")
                col1, col2 = st.columns(2)
                with col1:
                    st.info("""
                    **Configuración Local**
                    - ✓ Archivo .env existe
                    - ❓ OPENAI_API_KEY válida
                    - ❓ Conexión a internet
                    """)
                with col2:
                    st.info("""
                    **Streamlit Cloud**
                    - ❓ Secrets configurados
                    - ❓ Variables de entorno
                    - ❓ API key válida
                    """)
                
                if st.button("🔄 Reintentar Conexión IA", key="retry_ai"):
                    # Force reinitialization
                    for key in ['contenedor', 'caso_uso_maestro', 'ai_config_manager']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.experimental_rerun()
                
                st.stop()  # Don't proceed with analysis
                
        except Exception as e:
            logger.warning(f"Could not verify IA analyzer availability: {e}")
            # Continue anyway - let the analysis attempt fail gracefully
                
        from src.application.use_cases.analizar_excel_maestro_caso_uso import ComandoAnalisisExcelMaestro
        
        comando = ComandoAnalisisExcelMaestro(
            archivo_cargado=uploaded_file,
            nombre_archivo=uploaded_file.name,
            limpiar_repositorio=True
        )
        
        # Initialize progress display for real-time updates
        st.info("🤖 Iniciando análisis con Inteligencia Artificial...")
        
        # Display live progress fragment (updates every 0.5s automatically)
        live_batch_progress()
        
        # Execute analysis with progress callback
        resultado = caso_uso_maestro.ejecutar(comando)
        
        # Clear progress state after analysis
        if 'batch_progress_data' in st.session_state:
            del st.session_state.batch_progress_data
        
        if resultado.es_exitoso():
            # Memory management: cleanup previous analysis before storing new one
            _cleanup_previous_analysis()
            
            # PROGRESS TRACKING: Clear progress state on success
            if 'batch_progress_data' in st.session_state:
                del st.session_state.batch_progress_data
            
            st.session_state.analysis_results = resultado
            st.session_state.analysis_type = "maestro_ia"
            st.success("Análisis IA completado!")
            st.balloons()
            st.rerun()
        else:
            # PROGRESS TRACKING: Clear progress state on failure
            if 'batch_progress_data' in st.session_state:
                del st.session_state.batch_progress_data
            
            st.error(f"Error en análisis IA: {resultado.mensaje}")
            
    except ArchivoException as e:
        # PROGRESS TRACKING: Clear progress state on error
        if 'batch_progress_data' in st.session_state:
            del st.session_state.batch_progress_data
        st.error(f"Error procesando archivo: {str(e)}")
    except IAException as e:
        # PROGRESS TRACKING: Clear progress state on error
        if 'batch_progress_data' in st.session_state:
            del st.session_state.batch_progress_data
        st.error(f"Error de servicio IA: {str(e)}")
        st.info("Verifica que tu OpenAI API key esté configurada correctamente.")
    except Exception as e:
        # PROGRESS TRACKING: Clear progress state on error
        if 'batch_progress_data' in st.session_state:
            del st.session_state.batch_progress_data
        st.error(f"Error inesperado: {str(e)}")
        st.error("Este es un error no manejado. Por favor contacta soporte técnico.")
    finally:
        # PROGRESS TRACKING: Final cleanup guarantee
        st.session_state.ai_analysis_in_progress = False
        if 'batch_progress_data' in st.session_state:
            del st.session_state.batch_progress_data


def _create_comprehensive_emotions_chart(emociones_predominantes):
    """
    Create comprehensive emotion distribution chart - FIRST DISPLAY
    Shows detailed breakdown of all 16 granular emotions with intensities
    This is the primary emotion visualization showing rich emotional insights
    """
    if not emociones_predominantes or not isinstance(emociones_predominantes, dict):
        return None
        
    # Get all emotions sorted by intensity (highest first)
    sorted_emotions = sorted(emociones_predominantes.items(), key=lambda x: x[1], reverse=True)
    
    # Take all available emotions (up to 16)
    emotions = [emo for emo, intensity in sorted_emotions if intensity > 0][:16]
    intensities = [intensity for emo, intensity in sorted_emotions if intensity > 0][:16]
    
    if not emotions:
        return None
    
    # POLISH-002 FIX: Use constants for color mapping
    if CONSTANTS_AVAILABLE:
        emotion_colors = AIEngineConstants.EMOTION_COLORS
        default_color = AIEngineConstants.DEFAULT_EMOTION_COLOR
    else:
        # Fallback color mapping
        emotion_colors = {
            'satisfaccion': '#10B981', 'alegria': '#06D6A0', 'entusiasmo': '#FFD23F',
            'gratitud': '#118AB2', 'confianza': '#073B4C', 'frustracion': '#EF4444',
            'enojo': '#DC2626', 'decepcion': '#991B1B', 'preocupacion': '#F97316',
            'irritacion': '#EA580C', 'ansiedad': '#C2410C', 'tristeza': '#7C2D12',
            'confusion': '#6B7280', 'esperanza': '#8B5CF6', 'curiosidad': '#A855F7',
            'impaciencia': '#9333EA', 'neutral': '#9CA3AF'
        }
        default_color = '#8B5CF6'
    
    colors = [emotion_colors.get(emotion, default_color) for emotion in emotions]
    
    # Create horizontal bar chart for better readability with many emotions
    fig = go.Figure(data=[go.Bar(
        x=intensities,
        y=emotions,
        orientation='h',
        marker=dict(
            color=colors,
            opacity=0.8,
            line=dict(color='rgba(255,255,255,0.2)', width=1)
        ),
        text=[f'{intensity:.2f}' for intensity in intensities],
        textposition='auto',
        textfont=dict(color='white', size=11)
    )])
    
    fig.update_layout(
        title=dict(
            text="📊 Distribución Completa de Emociones Granulares",
            font=dict(size=16, color='white')
        ),
        xaxis_title="Intensidad",
        yaxis_title="Emociones",
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=AIEngineConstants.calculate_dynamic_chart_height(len(emotions)) if CONSTANTS_AVAILABLE else max(400, len(emotions) * 25 + 100),
        margin=dict(l=120, r=50, t=80, b=50),  # More left margin for emotion names
        xaxis=dict(
            gridcolor='rgba(255,255,255,0.1)',
            showgrid=True
        ),
        yaxis=dict(
            gridcolor='rgba(255,255,255,0.1)',
            showgrid=False
        )
    )
    
    return fig


def _create_sentiment_distribution_chart(distribucion_sentimientos):
    """Create pie chart for sentiment distribution"""
    if not distribucion_sentimientos:
        return None
        
    # Prepare data - handle both formats (legacy and new abbreviated)
    sentiments = ['Positivos', 'Neutrales', 'Negativos'] 
    values = [
        distribucion_sentimientos.get('positivo', distribucion_sentimientos.get('pos', 0)),
        distribucion_sentimientos.get('neutral', distribucion_sentimientos.get('neu', 0)), 
        distribucion_sentimientos.get('negativo', distribucion_sentimientos.get('neg', 0))
    ]
    
    colors = ['#10B981', '#6B7280', '#EF4444']  # Green, Gray, Red
    
    fig = go.Figure(data=[go.Pie(
        labels=sentiments,
        values=values,
        hole=0.3,
        marker_colors=colors,
        textinfo='label+percent',
        textfont_size=12
    )])
    
    fig.update_layout(
        title="Distribución de Sentimientos",
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=400
    )
    
    return fig


def _create_themes_chart(temas_relevantes):
    """Create horizontal bar chart for themes"""
    if not temas_relevantes:
        return None
    
    # Handle case where temas_relevantes might be a single theme from stats.tema_top
    if isinstance(temas_relevantes, dict) and len(temas_relevantes) == 1:
        # Single theme from abbreviated format
        themes = list(temas_relevantes.keys())
        relevances = list(temas_relevantes.values())
    else:
        # Multiple themes - take top 10
        themes = list(temas_relevantes.keys())[:10]
        relevances = list(temas_relevantes.values())[:10]
        
    # Ensure we have data to display
    if not themes or not relevances:
        return None
    
    fig = go.Figure(data=[go.Bar(
        x=relevances,
        y=themes,
        orientation='h',
        marker_color='#8B5CF6',  # Purple
        text=[f'{r:.1f}' for r in relevances],
        textposition='auto'
    )])
    
    fig.update_layout(
        title="Temas Más Relevantes",
        xaxis_title="Relevancia",
        yaxis_title="Temas", 
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=400
    )
    
    return fig


def _create_emotions_donut_chart(emociones_predominantes):
    """Create donut chart for emotions with intensity"""
    if not emociones_predominantes:
        return None
        
    emotions = list(emociones_predominantes.keys())[:8]  # Top 8 emotions
    intensities = list(emociones_predominantes.values())[:8]
    
    # Color mapping for emotions
    emotion_colors = {
        'satisfaccion': '#10B981',
        'frustracion': '#EF4444', 
        'enojo': '#DC2626',
        'alegria': '#F59E0B',
        'decepcion': '#6B7280',
        'preocupacion': '#F97316',
        'neutral': '#9CA3AF'
    }
    
    colors = [emotion_colors.get(e, '#8B5CF6') for e in emotions]
    
    fig = go.Figure(data=[go.Pie(
        labels=emotions,
        values=intensities,
        hole=0.4,
        marker_colors=colors,
        textinfo='label+value',
        textfont_size=11
    )])
    
    fig.update_layout(
        title="Emociones Detectadas por IA", 
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=400
    )
    
    return fig


def _create_token_usage_gauge(tokens_utilizados, max_tokens=8000):
    """Create gauge chart for token usage"""
    # Validate input data
    tokens_used = tokens_utilizados if tokens_utilizados and tokens_utilizados > 0 else 0
    percentage = (tokens_used / max_tokens) * 100
    
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=tokens_used,
        delta={'reference': max_tokens * 0.8},  # 80% reference
        gauge={
            'axis': {'range': [None, max_tokens]},
            'bar': {'color': "#06B6D4"},  # Cyan
            'steps': [
                {'range': [0, max_tokens * 0.6], 'color': "rgba(16, 185, 129, 0.2)"},  # Green zone
                {'range': [max_tokens * 0.6, max_tokens * 0.8], 'color': "rgba(245, 158, 11, 0.2)"},  # Yellow zone
                {'range': [max_tokens * 0.8, max_tokens], 'color': "rgba(239, 68, 68, 0.2)"}  # Red zone
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': max_tokens * 0.9
            }
        }
    ))
    
    fig.update_layout(
        title=f"Uso de Tokens ({percentage:.1f}%)",
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)', 
        height=300
    )
    
    return fig


def _create_confidence_histogram(comentarios_analizados):
    """Create histogram for confidence distribution"""
    if not comentarios_analizados:
        return None
        
    # Extract confidence values - handle both DTO objects and dict responses
    confidences = []
    for comentario in comentarios_analizados:
        if isinstance(comentario, dict):
            # Handle abbreviated format from new AI response: 'conf' or legacy 'confianza'
            conf = comentario.get('conf', comentario.get('confianza', 0.5))
            confidences.append(float(conf))
        elif hasattr(comentario, 'confianza_general'):
            # Handle AnalisisComentario entity objects
            confidences.append(float(comentario.confianza_general))
        else:
            # Fallback for unknown format
            confidences.append(0.5)
    
    if not confidences:
        return None
        
    fig = go.Figure(data=[go.Histogram(
        x=confidences,
        nbinsx=10,
        marker_color='#8B5CF6',
        opacity=0.7
    )])
    
    fig.update_layout(
        title="Distribución de Confianza del Análisis IA",
        xaxis_title="Confianza",
        yaxis_title="Frecuencia", 
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=300
    )
    
    return fig


def _create_batch_processing_timeline(analisis):
    """Create timeline chart for batch processing metrics"""
    if not analisis or not hasattr(analisis, 'tiempo_analisis'):
        return None
    
    # Estimate batch information
    total_comments = analisis.total_comentarios
    batch_size = 50  # CONSISTENCY: Updated to match optimized batch size
    num_batches = max(1, (total_comments + batch_size - 1) // batch_size)  # Ceiling division
    time_per_batch = analisis.tiempo_analisis / num_batches if num_batches > 0 else 0
    
    # Create timeline data
    batch_numbers = list(range(1, num_batches + 1))
    cumulative_times = [i * time_per_batch for i in batch_numbers]
    
    fig = go.Figure()
    
    # Add timeline line
    fig.add_trace(go.Scatter(
        x=batch_numbers,
        y=cumulative_times,
        mode='lines+markers',
        name='Tiempo Acumulado',
        line=dict(color='#06B6D4', width=3),
        marker=dict(size=8)
    ))
    
    fig.update_layout(
        title=f"Timeline de Procesamiento ({num_batches} lotes)",
        xaxis_title="Número de Lote",
        yaxis_title="Tiempo Acumulado (segundos)",
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        height=300
    )
    
    return fig


def _create_ai_metrics_summary(analisis):
    """Create summary metrics visualization"""
    if not analisis:
        return None
    
    # Create multi-metric gauge chart
    fig = make_subplots(
        rows=1, cols=3,
        specs=[[{'type': 'indicator'}, {'type': 'indicator'}, {'type': 'indicator'}]],
        subplot_titles=('Confianza IA', 'Eficiencia Tokens', 'Velocidad Análisis')
    )
    
    # Confidence gauge
    confidence_pct = analisis.confianza_general * 100 if analisis.confianza_general else 50
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=confidence_pct,
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': "#10B981"},
            'steps': [
                {'range': [0, 60], 'color': "rgba(239, 68, 68, 0.2)"},
                {'range': [60, 80], 'color': "rgba(245, 158, 11, 0.2)"},
                {'range': [80, 100], 'color': "rgba(16, 185, 129, 0.2)"}
            ]
        },
        number={'suffix': '%'}
    ), row=1, col=1)
    
    # Token efficiency gauge
    max_tokens = 8000
    token_efficiency = ((max_tokens - analisis.tokens_utilizados) / max_tokens) * 100 if analisis.tokens_utilizados else 50
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=token_efficiency,
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': "#8B5CF6"},
        },
        number={'suffix': '%'}
    ), row=1, col=2)
    
    # Processing speed gauge (comments per minute)
    total_comments = analisis.total_comentarios if analisis.total_comentarios else 0
    analysis_time = analisis.tiempo_analisis if analisis.tiempo_analisis else 1
    speed = (total_comments / analysis_time * 60) if analysis_time > 0 else 0
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=speed,
        gauge={
            'axis': {'range': [0, 200]},
            'bar': {'color': "#06B6D4"},
        },
        number={'suffix': '/min'}
    ), row=1, col=3)
    
    fig.update_layout(
        font=dict(color='white'),
        paper_bgcolor='rgba(0,0,0,0)',
        height=250
    )
    
    return fig


def _create_professional_excel(resultado):
    """Create comprehensive Excel export from IA analysis using real DTO structure"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis IA Completo"
    
    # Styling
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    
    # Header section
    ws['A1'] = "Personal Paraguay - Análisis con Inteligencia Artificial"
    ws['A1'].font = header_font
    ws['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = "Método: AnalizadorMaestroIA + GPT-4"
    
    if hasattr(resultado, 'analisis_completo_ia') and resultado.analisis_completo_ia:
        analisis = resultado.analisis_completo_ia
        
        # Executive Summary
        ws['A5'] = "RESUMEN EJECUTIVO IA"
        ws['A5'].font = section_font
        ws['A6'] = f"Total comentarios analizados: {analisis.total_comentarios}"
        ws['A7'] = f"Tendencia general: {analisis.tendencia_general}"
        ws['A8'] = f"Confianza del análisis: {analisis.confianza_general:.1f}%"
        ws['A9'] = f"Modelo IA utilizado: {analisis.modelo_utilizado}"
        ws['A10'] = f"Tiempo de procesamiento: {analisis.tiempo_analisis:.1f}s"
        ws['A11'] = f"Tokens consumidos: {analisis.tokens_utilizados:,}"
        
        # IA Narrative Summary
        ws['A13'] = "ANÁLISIS NARRATIVO IA"
        ws['A13'].font = section_font
        ws.merge_cells('A14:E14')
        ws['A14'] = analisis.resumen_ejecutivo
        ws['A14'].alignment = Alignment(wrap_text=True)
        
        # Sentiment distribution
        ws['A16'] = "DISTRIBUCIÓN DE SENTIMIENTOS"
        ws['A16'].font = section_font
        row = 17
        for sentimiento, cantidad in analisis.distribucion_sentimientos.items():
            ws[f'A{row}'] = sentimiento
            ws[f'B{row}'] = cantidad
            ws[f'C{row}'] = f"{(cantidad/analisis.total_comentarios)*100:.1f}%"
            row += 1
        
        # Top themes with relevance
        ws[f'A{row + 1}'] = "TEMAS MÁS RELEVANTES"
        ws[f'A{row + 1}'].font = section_font
        row += 2
        for tema, relevancia in list(analisis.temas_mas_relevantes.items())[:8]:
            ws[f'A{row}'] = tema
            ws[f'B{row}'] = f"{relevancia:.2f}"
            ws[f'C{row}'] = "Alta" if relevancia > 0.7 else "Media" if relevancia > 0.4 else "Baja"
            row += 1
        
        # ENHANCED: Comprehensive Emotion Distribution with Statistics
        ws[f'A{row + 1}'] = "DISTRIBUCIÓN COMPLETA DE EMOCIONES GRANULARES"
        ws[f'A{row + 1}'].font = section_font  
        row += 2
        
        # Column headers for emotion statistics
        ws[f'A{row}'] = "Emoción"
        ws[f'B{row}'] = "Intensidad"
        ws[f'C{row}'] = "Porcentaje"
        ws[f'D{row}'] = "Clasificación"
        ws[f'E{row}'] = "Tipo"
        
        # Style headers
        header_row = row
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{header_row}'].font = Font(bold=True)
        
        row += 1
        
        # Sort emotions by intensity (highest first) and show ALL emotions
        if analisis.emociones_predominantes:
            sorted_emotions = sorted(analisis.emociones_predominantes.items(), 
                                   key=lambda x: x[1], reverse=True)
            
            # Define emotion types for categorization
            emotion_types = {
                'satisfaccion': 'Positiva', 'alegria': 'Positiva', 'entusiasmo': 'Positiva', 
                'gratitud': 'Positiva', 'confianza': 'Positiva',
                'frustracion': 'Negativa', 'enojo': 'Negativa', 'decepcion': 'Negativa',
                'preocupacion': 'Negativa', 'irritacion': 'Negativa', 'ansiedad': 'Negativa',
                'tristeza': 'Negativa', 'confusion': 'Neutra', 'esperanza': 'Neutra',
                'curiosidad': 'Neutra', 'impaciencia': 'Neutra', 'neutral': 'Neutra'
            }
            
            total_intensity = sum(analisis.emociones_predominantes.values())
            
            for emocion, intensidad in sorted_emotions:
                # Calculate percentage of total emotional expression
                percentage = (intensidad / total_intensity * 100) if total_intensity > 0 else 0
                
                # POLISH-002 FIX: Use constants for intensity classification
                if CONSTANTS_AVAILABLE:
                    clasificacion = AIEngineConstants.classify_emotion_intensity(intensidad)
                else:
                    # Fallback classification
                    if intensidad >= 0.8:
                        clasificacion = "Muy Intensa"
                    elif intensidad >= 0.6:
                        clasificacion = "Intensa"
                    elif intensidad >= 0.4:
                        clasificacion = "Moderada"
                    elif intensidad >= 0.2:
                        clasificacion = "Leve"
                    else:
                        clasificacion = "Muy Leve"
                
                # Get emotion type
                tipo = emotion_types.get(emocion, 'Desconocida')
                
                # Write to Excel
                ws[f'A{row}'] = emocion.replace('_', ' ').title()
                ws[f'B{row}'] = f"{intensidad:.3f}"
                ws[f'C{row}'] = f"{percentage:.1f}%"
                ws[f'D{row}'] = clasificacion
                ws[f'E{row}'] = tipo
                row += 1
            
            # Add summary statistics
            row += 1
            ws[f'A{row}'] = "ESTADÍSTICAS DE EMOCIONES"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Calculate emotion type distributions
            tipo_counts = {'Positiva': 0, 'Negativa': 0, 'Neutra': 0}
            for emocion, intensidad in analisis.emociones_predominantes.items():
                tipo = emotion_types.get(emocion, 'Neutra')
                tipo_counts[tipo] += intensidad
            
            for tipo, total_intensity in tipo_counts.items():
                ws[f'A{row}'] = f"Total {tipo}s"
                ws[f'B{row}'] = f"{total_intensity:.2f}"
                ws[f'C{row}'] = f"{(total_intensity/sum(tipo_counts.values())*100):.1f}%" if sum(tipo_counts.values()) > 0 else "0%"
                row += 1
        
        # Pain points with severity
        if analisis.dolores_mas_severos:
            ws[f'A{row + 1}'] = "PUNTOS DE DOLOR CRÍTICOS"
            ws[f'A{row + 1}'].font = section_font
            row += 2
            for dolor, severidad in list(analisis.dolores_mas_severos.items())[:5]:
                ws[f'A{row}'] = dolor
                ws[f'B{row}'] = f"{severidad:.1f}"
                ws[f'C{row}'] = "Crítico" if severidad > 8 else "Alto" if severidad > 6 else "Medio"
                row += 1
        
        # IA Recommendations
        ws[f'A{row + 1}'] = "RECOMENDACIONES ACCIONABLES IA"
        ws[f'A{row + 1}'].font = section_font
        row += 2
        for i, recomendacion in enumerate(analisis.recomendaciones_principales, 1):
            ws[f'A{row}'] = f"Recomendación {i}"
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = recomendacion
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
            
    else:
        # Fallback structure
        ws['A5'] = "DATOS LIMITADOS DISPONIBLES"
        ws['A6'] = f"Total comentarios: {getattr(resultado, 'total_comentarios', 0)}"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _cleanup_previous_analysis():
    """
    Limpia análisis previos de session state para prevenir acumulación de memoria
    """
    cleanup_keys = [
        'analysis_results',
        'analysis_type'
    ]
    
    for key in cleanup_keys:
        if key in st.session_state:
            # Clear large objects to free memory
            del st.session_state[key]
    
    # Also cleanup repository cache if available
    if 'contenedor' in st.session_state and st.session_state.contenedor:
        try:
            repo = st.session_state.contenedor.obtener_repositorio_comentarios()
            if hasattr(repo, 'limpiar'):
                repo.limpiar()
        except Exception:
            pass  # Ignore errors in cleanup
    
    # Force garbage collection
    import gc
    gc.collect()


st.title("Subir y Analizar Comentarios")

st.markdown("""
Sube tu archivo Excel o CSV con comentarios de clientes para análisis automático.
""")

# File upload
st.markdown("### Cargar Archivo")

uploaded_file = st.file_uploader(
    "Selecciona tu archivo",
    type=['xlsx', 'xls', 'csv'],
    help="Formatos soportados: Excel (.xlsx, .xls) y CSV"
)

if uploaded_file:
    # Basic file validation
    file_size_mb = uploaded_file.size / (1024 * 1024)
    
    if file_size_mb > 5:
        st.error("Archivo muy grande. Máximo 5MB.")
        st.stop()
    
    st.success(f"Archivo válido: {uploaded_file.name} ({file_size_mb:.1f}MB)")
    
    # Enhanced file preview with glassmorphism if available
    with st.expander("👀 Vista Previa del Archivo", expanded=True):
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file, nrows=5)
                df_full = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file, nrows=5)
                df_full = pd.read_excel(uploaded_file)
            
            # File stats with enhanced display
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            with col_stats1:
                st.metric("📊 Total Filas", len(df_full))
            with col_stats2:
                st.metric("📋 Columnas", len(df_full.columns))
            with col_stats3:
                # Detect comment column
                comment_cols = ['comentario', 'comment', 'comentarios', 'feedback', 'review']
                comment_col = None
                for col in df_full.columns:
                    if any(cc in col.lower() for cc in comment_cols):
                        comment_col = col
                        break
                st.metric("💬 Comentarios", len(df_full[comment_col].dropna()) if comment_col else "No detectados")
            
            # Data preview
            st.markdown("**Primeras 5 filas:**")
            st.dataframe(df, width="stretch")
            
            # Column analysis
            if comment_col:
                st.success(f"✅ Columna de comentarios detectada: **{comment_col}**")
            else:
                st.warning("⚠️ No se detectó columna de comentarios clara")
                
        except Exception as e:
            st.warning(f"No se pudo generar vista previa: {str(e)}")
    
    # Analysis section
    st.markdown("### Análisis")
    
    # Check if IA system is ready using validator for consistency
    try:
        from src.presentation.streamlit.session_validator import is_ia_system_ready
        if not is_ia_system_ready():
            st.error("Sistema IA no inicializado. Recarga la página o verifica configuración OpenAI.")
            st.info("💡 Asegúrate de que la API key de OpenAI esté configurada correctamente.")
            st.stop()
    except ImportError:
        # Fallback to direct check if validator not available
        if 'caso_uso_maestro' not in st.session_state or not st.session_state.caso_uso_maestro:
            st.error("Sistema IA no inicializado. Recarga la página o verifica configuración OpenAI.")
            st.stop()
    
    # IA Analysis (single button - pure IA app)
    if st.button("Analizar con Inteligencia Artificial", type="primary", width="stretch"):
        _run_analysis(uploaded_file, "ai")

# Results section
if 'analysis_results' in st.session_state:
    st.markdown("---")
    st.markdown("### Resultados")
    
    results = st.session_state.analysis_results
    analysis_type = st.session_state.get('analysis_type', 'unknown')
    
    # IA Analysis status (pure IA app)
    st.success("Análisis con Inteligencia Artificial completado")
    
    # Show IA analysis results (pure IA format)
    if hasattr(results, 'es_exitoso') and results.es_exitoso():
        # Enhanced IA Analysis with Interactive Charts
        if hasattr(results, 'analisis_completo_ia') and results.analisis_completo_ia:
            analisis = results.analisis_completo_ia
            
            st.markdown("---")  # Visual separator
            
            # Key Metrics Row
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Comentarios", analisis.total_comentarios)
            with col2:
                st.metric("Tiempo IA", f"{analisis.tiempo_analisis:.1f}s")
            with col3:
                sentiments = analisis.distribucion_sentimientos
                # Handle multiple format possibilities: 'positivo', 'pos', 'POSITIVO'
                positivos = sentiments.get('positivo', sentiments.get('pos', sentiments.get('POSITIVO', 0)))
                st.metric("Positivos", positivos)
            with col4:
                # Handle multiple format possibilities: 'negativo', 'neg', 'NEGATIVO'  
                negativos = sentiments.get('negativo', sentiments.get('neg', sentiments.get('NEGATIVO', 0)))
                st.metric("Negativos", negativos)
            
            # AI Metrics Summary Gauges
            ai_metrics_chart = _create_ai_metrics_summary(analisis)
            if ai_metrics_chart:
                st.plotly_chart(ai_metrics_chart, use_container_width=True)
        else:
            # Fallback metrics if IA structure incomplete
            with col1:
                st.metric("Total Comentarios", results.total_comentarios)
            with col2:
                st.metric("Estado", "Procesado")
            with col3:
                st.metric("Método", "IA Avanzada")
            with col4:
                st.metric("Calidad", "Máxima")
                
            # Fallback message for incomplete data
            st.info("📊 Visualizaciones avanzadas disponibles con análisis IA completo")
            
        # IA Insights (pure mechanical mapping using REAL DTO structure)
        st.markdown("#### Insights de Inteligencia Artificial")
        
        if hasattr(results, 'analisis_completo_ia') and results.analisis_completo_ia:
            analisis = results.analisis_completo_ia
            
            # Display IA executive summary first
            st.markdown("**Resumen Ejecutivo (Generado por IA):**")
            st.info(analisis.resumen_ejecutivo)
            
            # IA metrics in columns
            col_ia1, col_ia2 = st.columns(2)
            
            with col_ia1:
                st.markdown("**📊 Métricas IA:**")
                st.metric("Confianza General", f"{analisis.confianza_general:.1f}%")
                st.metric("Modelo Utilizado", analisis.modelo_utilizado)
                st.metric("Tokens Utilizados", f"{analisis.tokens_utilizados:,}")
            
            with col_ia2:
                st.markdown("**🎯 Análisis:**")
                st.metric("Tendencia General", analisis.tendencia_general.title())
                
                # Count critical comments from individual analysis
                if hasattr(results, 'comentarios_analizados') and results.comentarios_analizados:
                    criticos = len([c for c in results.comentarios_analizados if hasattr(c, 'es_critico') and c.es_critico()])
                    st.metric("Comentarios Críticos", criticos)
            
            # ENHANCED VISUALIZATION: AI Analysis Charts
            st.markdown("#### 📊 Visualización de Análisis IA")
            
            # FIRST CHART: Comprehensive Emotion Distribution (Most Important)
            if analisis.emociones_predominantes:
                st.markdown("##### 🎭 Distribución Completa de Emociones Detectadas")
                emotions_main_chart = _create_comprehensive_emotions_chart(analisis.emociones_predominantes)
                if emotions_main_chart:
                    st.plotly_chart(emotions_main_chart, use_container_width=True)
                else:
                    st.info("📊 No se detectaron emociones suficientes para visualización")
            
            # Additional Charts in columns
            st.markdown("##### 📊 Métricas Adicionales")
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                # Sentiment Distribution Chart
                if analisis.distribucion_sentimientos:
                    sentiment_chart = _create_sentiment_distribution_chart(analisis.distribucion_sentimientos)
                    if sentiment_chart:
                        st.plotly_chart(sentiment_chart, use_container_width=True)
                
                # Token Usage Gauge
                if analisis.tokens_utilizados:
                    token_chart = _create_token_usage_gauge(analisis.tokens_utilizados)
                    if token_chart:
                        st.plotly_chart(token_chart, use_container_width=True)
            
            with col_chart2:
                # Themes Chart
                if analisis.temas_mas_relevantes:
                    themes_chart = _create_themes_chart(analisis.temas_mas_relevantes)
                    if themes_chart:
                        st.plotly_chart(themes_chart, use_container_width=True)
                
                # Emotions Chart  
                if analisis.emociones_predominantes:
                    emotions_chart = _create_emotions_donut_chart(analisis.emociones_predominantes)
                    if emotions_chart:
                        st.plotly_chart(emotions_chart, use_container_width=True)
            
            # Additional Insights Charts
            col_insight1, col_insight2 = st.columns(2) 
            
            with col_insight1:
                # Confidence Distribution Chart
                if hasattr(results, 'comentarios_analizados') and results.comentarios_analizados:
                    confidence_chart = _create_confidence_histogram(results.comentarios_analizados)
                    if confidence_chart:
                        st.plotly_chart(confidence_chart, use_container_width=True)
            
            with col_insight2:
                # Batch Processing Timeline
                batch_timeline = _create_batch_processing_timeline(analisis)
                if batch_timeline:
                    st.plotly_chart(batch_timeline, use_container_width=True)
            
            # Text Summary (Reduced, complementing charts)
            if analisis.temas_mas_relevantes:
                st.markdown("**🏷️ Top 3 Temas Detectados:**")
                for tema, relevancia in list(analisis.temas_mas_relevantes.items())[:3]:
                    st.markdown(f"• **{tema}**: {relevancia:.1f}")
            
            if analisis.emociones_predominantes:
                st.markdown("**😊 Top 3 Emociones Identificadas:**") 
                for emocion, intensidad in list(analisis.emociones_predominantes.items())[:3]:
                    st.markdown(f"• **{emocion}**: {intensidad:.1f}")
            
            # Pain points from IA
            if analisis.dolores_mas_severos:
                st.markdown("**⚠️ Puntos de Dolor Más Severos (IA):**")
                for dolor, severidad in list(analisis.dolores_mas_severos.items())[:3]:
                    st.markdown(f"• **{dolor}**: {severidad:.1f} severidad")
            
            # IA Recommendations
            if analisis.recomendaciones_principales:
                st.markdown("**💡 Recomendaciones de IA:**")
                for i, recomendacion in enumerate(analisis.recomendaciones_principales[:3], 1):
                    st.markdown(f"{i}. {recomendacion}")
            
            # Critical comments detected by IA (using real data structure)
            if hasattr(results, 'comentarios_analizados') and results.comentarios_analizados:
                criticos_ia = [c for c in results.comentarios_analizados 
                              if hasattr(c, 'es_critico') and c.es_critico()]
                
                if criticos_ia:
                    with st.expander(f"🚨 {len(criticos_ia)} comentarios críticos (IA)"):
                        st.warning("Comentarios que requieren atención inmediata según análisis IA:")
                        for i, comentario in enumerate(criticos_ia[:5], 1):
                            st.warning(f"**{i}.** {comentario.texto_original}")
                            
                            # Show IA-detected urgency and recommendations
                            if hasattr(comentario, 'puntos_dolor') and comentario.puntos_dolor:
                                dolores_texto = ", ".join([p.contexto_especifico for p in comentario.puntos_dolor[:2] if p.contexto_especifico])
                                st.caption(f"🎯 Puntos de dolor IA: {dolores_texto}")
                            
                            if hasattr(comentario, 'recomendaciones') and comentario.recomendaciones:
                                st.caption(f"💡 Recomendación IA: {comentario.recomendaciones[0]}")
        else:
            st.info("Análisis IA completado - formato de datos simplificado")
        
        # Export IA results
        st.markdown("#### Exportar Análisis IA")
        try:
            # Generate Excel directly for download button
            with st.spinner("Preparando Excel profesional..."):
                excel_data = _create_professional_excel(results)
            
            # Direct download button - no extra step needed
            st.download_button(
                "📊 Descargar Excel Profesional", 
                excel_data,
                f"analisis_ia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            st.success("✅ Excel listo para descarga con análisis completo de IA")
        except Exception as e:
            st.error(f"❌ Error generando Excel: {str(e)}")
            logger.error(f"Error en generación de Excel: {str(e)}")
    else:
        st.error(f"Error en análisis IA: {results.mensaje if hasattr(results, 'mensaje') else 'Error desconocido'}")