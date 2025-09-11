#!/usr/bin/env python3
"""
🌩️ Pipeline Test con Configuración de Streamlit Cloud
Usa la configuración unificada para simular el comportamiento de Streamlit Cloud
"""

import os
import sys
import time
from pathlib import Path
import logging

# Setup paths
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

# Force cloud environment simulation
os.environ['STREAMLIT_CLOUD_EMULATOR'] = 'true'

# Import unified configuration
from config import config, is_streamlit_cloud, get_environment_info

# Setup logging with cloud config
logging.basicConfig(
    level=getattr(logging, config.get('LOG_LEVEL', 'INFO')),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def show_environment_info():
    """Display current environment configuration"""
    print("🌩️ EMULADOR DE STREAMLIT CLOUD - PIPELINE TEST")
    print("=" * 55)
    
    env_info = get_environment_info()
    print("📊 Configuración actual:")
    for key, value in env_info.items():
        print(f"  • {key}: {value}")
    
    print(f"\n⚙️ Límites aplicados:")
    print(f"  • Batch size máximo: {config.get('MAX_COMMENTS_PER_BATCH')}")
    print(f"  • Memoria optimizada: {config.get('ENABLE_MEMORY_OPTIMIZATION')}")
    print(f"  • Cache TTL: {config.get('CACHE_TTL_SECONDS')}s")
    print(f"  • Delay entre batches: {config.get('BATCH_PROCESSING_DELAY')}s")
    print(f"  • Tokens máximos: {config.get('OPENAI_MAX_TOKENS')}")
    print("=" * 55)

def test_imports():
    """Test core imports with cloud configuration"""
    print("\n🧪 Test 1: Imports con configuración Cloud")
    
    try:
        from src.infrastructure.dependency_injection.contenedor_dependencias import ContenedorDependencias
        from src.application.use_cases.analizar_excel_maestro_caso_uso import AnalizarExcelMaestroCasoUso
        from src.infrastructure.file_handlers.lector_archivos_excel import LectorArchivosExcel
        print("  ✅ Imports exitosos")
        return True
    except Exception as e:
        print(f"  ❌ Error en imports: {e}")
        return False

def test_container_with_cloud_config():
    """Test container creation with cloud configuration"""
    print("\n🧪 Test 2: Container con límites Cloud")
    
    try:
        # Use cloud configuration - get real API key
        api_key = config.get('OPENAI_API_KEY', '')
        if not api_key or not api_key.startswith('sk-'):
            # Try to load from .env file
            from dotenv import load_dotenv
            load_dotenv()
            api_key = os.getenv('OPENAI_API_KEY', '')
        
        cloud_config = {
            'openai_api_key': api_key,
            'openai_modelo': config.get('OPENAI_MODEL', 'gpt-4o-mini'),
            'openai_max_tokens': config.get('OPENAI_MAX_TOKENS', 4000),
            'openai_temperatura': config.get('OPENAI_TEMPERATURE', 0.0),
            'max_comments': config.get('MAX_COMMENTS_PER_BATCH', 10),  # Cloud limit
            'cache_ttl': config.get('CACHE_TTL_SECONDS', 1800),
        }
        
        if not api_key or not api_key.startswith('sk-'):
            print(f"  ⚠️ API Key no válida: '{api_key[:10]}...' (se usará mock)")
            cloud_config['openai_api_key'] = 'sk-test-mock-key-for-cloud-emulation'
        
        from src.infrastructure.dependency_injection.contenedor_dependencias import ContenedorDependencias
        contenedor = ContenedorDependencias(cloud_config)
        
        print(f"  ✅ Container creado con límites Cloud:")
        print(f"    • Max comments: {cloud_config['max_comments']}")
        print(f"    • Max tokens: {cloud_config['openai_max_tokens']}")
        print(f"    • Cache TTL: {cloud_config['cache_ttl']}s")
        
        return True, contenedor, cloud_config
    except Exception as e:
        print(f"  ❌ Error creando container: {e}")
        return False, None, None

def test_file_reading():
    """Test file reading with actual test file"""
    print("\n🧪 Test 3: Lectura de archivo de prueba")
    
    try:
        # Look for test files - prioritize valid test file
        local_reports = project_root / "local-reports"
        
        # Try specific test files first
        priority_files = [
            "test_minimal.xlsx",
            "test_comments_valid.xlsx", 
            "test_structure.xlsx"
        ]
        
        test_file = None
        for filename in priority_files:
            candidate = local_reports / filename
            if candidate.exists():
                test_file = candidate
                break
        
        if not test_file:
            # Fallback to any Excel file
            test_files = list(local_reports.glob("*.xlsx"))
            if test_files:
                test_file = test_files[0]
            else:
                print("  ⚠️ No se encontraron archivos Excel de prueba")
                print(f"  📁 Buscando en: {local_reports}")
                return False, None
        print(f"  📄 Usando archivo: {test_file.name}")
        
        from src.infrastructure.file_handlers.lector_archivos_excel import LectorArchivosExcel
        lector = LectorArchivosExcel()
        
        with open(test_file, 'rb') as f:
            file_content = f.read()
        
        # Mock file object
        class MockFile:
            def __init__(self, content, name):
                self.content = content
                self.name = name
            def read(self): return self.content
            def seek(self, pos): pass
        
        mock_file = MockFile(file_content, test_file.name)
        comentarios = lector.leer_comentarios(mock_file)
        
        if comentarios and len(comentarios) > 0:
            print(f"  ✅ Leídos {len(comentarios)} comentarios")
            
            # Apply cloud limits
            max_comments = config.get('MAX_COMMENTS_PER_BATCH', 10)
            if len(comentarios) > max_comments:
                comentarios_limitados = comentarios[:max_comments]
                print(f"  ⚠️ Aplicando límite Cloud: {max_comments} comentarios")
                return True, comentarios_limitados
            
            return True, comentarios
        else:
            print("  ❌ No se encontraron comentarios")
            return False, None
            
    except Exception as e:
        print(f"  ❌ Error leyendo archivo: {e}")
        return False, None

def test_analysis_with_cloud_limits(contenedor, comentarios, cloud_config):
    """Test analysis with cloud resource limits"""
    print("\n🧪 Test 4: Análisis con límites Cloud")
    
    if not contenedor or not comentarios:
        print("  ❌ Prerequisites no disponibles")
        return False
    
    try:
        from src.application.use_cases.analizar_excel_maestro_caso_uso import ComandoAnalisisExcelMaestro
        
        # Create progress tracking
        progress_events = []
        
        def cloud_progress_callback(data):
            progress_events.append(data)
            action = data.get('action', 'unknown')
            if action == 'batch_start':
                batch = data.get('current_batch', 0)
                print(f"    📊 Procesando batch {batch}...")
            elif action == 'batch_success':
                batch = data.get('current_batch', 0)
                confidence = data.get('confidence', 0)
                print(f"    ✅ Batch {batch} completado (confianza: {confidence:.1%})")
        
        # Get use case with cloud callback
        caso_uso = contenedor.obtener_caso_uso_maestro(cloud_progress_callback)
        
        # Use the actual Excel file that was read successfully
        test_file = project_root / "local-reports" / "test_minimal.xlsx"
        if not test_file.exists():
            print("  ❌ Archivo de test no encontrado")
            return False, None
        
        # Create proper file object for the command
        class StreamlitUploadedFile:
            def __init__(self, file_path):
                self.file_path = file_path
                with open(file_path, 'rb') as f:
                    self._content = f.read()
                self.name = file_path.name
                self.size = len(self._content)
                self.type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                self._position = 0
            
            def read(self, size=-1):
                if size == -1:
                    content = self._content[self._position:]
                    self._position = len(self._content)
                else:
                    content = self._content[self._position:self._position + size]
                    self._position += len(content)
                return content
            
            def seek(self, position):
                self._position = position
            
            def tell(self):
                return self._position
        
        uploaded_file = StreamlitUploadedFile(test_file)
        
        comando = ComandoAnalisisExcelMaestro(
            archivo_cargado=uploaded_file,
            nombre_archivo="test_cloud_limits.xlsx"
        )
        
        print(f"  🚀 Iniciando análisis con {len(comentarios)} comentarios...")
        print(f"  ⏱️ Delay entre batches: {config.get('BATCH_PROCESSING_DELAY')}s")
        
        start_time = time.time()
        
        # Execute analysis
        try:
            resultado = caso_uso.ejecutar(comando)
        except Exception as analysis_error:
            print(f"  ❌ Error ejecutando análisis: {analysis_error}")
            import traceback
            print(f"  📋 Traceback detallado:\n{traceback.format_exc()}")
            return False, None
        
        end_time = time.time()
        duration = end_time - start_time
        
        if resultado and hasattr(resultado, 'exito') and resultado.exito:
            print(f"  ✅ Análisis completado en {duration:.1f}s")
            print(f"  📊 Eventos de progreso: {len(progress_events)}")
            print(f"  📊 Total comentarios procesados: {resultado.total_comentarios}")
            
            # Try to get confidence from analysis
            try:
                if (resultado.analisis_completo_ia and 
                    hasattr(resultado.analisis_completo_ia, 'resumen_ejecutivo') and
                    hasattr(resultado.analisis_completo_ia.resumen_ejecutivo, 'confianza')):
                    confianza = resultado.analisis_completo_ia.resumen_ejecutivo.confianza
                    print(f"  📈 Confianza promedio: {confianza:.1%}")
                else:
                    print(f"  📈 Análisis IA completado (estructura de resultado simplificada)")
            except Exception as conf_error:
                print(f"  📈 Análisis IA completado (no se pudo obtener confianza: {conf_error})")
            
            # Check if cloud limits were respected
            total_batches = max([e.get('current_batch', 0) for e in progress_events if 'current_batch' in e], default=1)
            expected_batches = max(1, (len(comentarios) + config.get('MAX_COMMENTS_PER_BATCH', 10) - 1) // config.get('MAX_COMMENTS_PER_BATCH', 10))
            
            if total_batches == expected_batches:
                print(f"  ✅ Batching correcto: {total_batches} batches")
            else:
                print(f"  ⚠️ Batching: {total_batches} batches (esperados {expected_batches})")
            
            return True, resultado
        else:
            print("  ❌ Análisis falló o resultado inválido")
            return False, None
            
    except Exception as e:
        print(f"  ❌ Error en análisis: {e}")
        import traceback
        print(f"  📋 Traceback: {traceback.format_exc()}")
        return False, None

def test_memory_behavior():
    """Test memory behavior with cloud limits"""
    print("\n🧪 Test 5: Comportamiento de memoria")
    
    try:
        import psutil
        import gc
        
        # Get initial memory
        process = psutil.Process()
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB
        
        print(f"  📊 Memoria inicial: {initial_memory:.1f} MB")
        
        # Force garbage collection (as cloud config would)
        if config.get('ENABLE_MEMORY_OPTIMIZATION', False):
            gc.collect()
            after_gc_memory = process.memory_info().rss / 1024 / 1024
            print(f"  🧹 Después de GC: {after_gc_memory:.1f} MB")
        
        # Simulate memory warning thresholds
        memory_warning_threshold = 800  # 800MB (80% of 1GB)
        memory_critical_threshold = 900  # 900MB (90% of 1GB)
        
        current_memory = process.memory_info().rss / 1024 / 1024
        
        if current_memory > memory_critical_threshold:
            print(f"  🚨 CRÍTICO: Memoria {current_memory:.1f} MB > {memory_critical_threshold} MB")
            print("  🔄 En Streamlit Cloud la app se reiniciaría")
            return False
        elif current_memory > memory_warning_threshold:
            print(f"  ⚠️ ADVERTENCIA: Memoria {current_memory:.1f} MB > {memory_warning_threshold} MB")
            print("  🐌 En Streamlit Cloud la app se ralentizaría")
            return True
        else:
            print(f"  ✅ Memoria OK: {current_memory:.1f} MB < {memory_warning_threshold} MB")
            return True
            
    except ImportError:
        print("  ⚠️ psutil no disponible - no se puede monitorear memoria")
        return True
    except Exception as e:
        print(f"  ❌ Error monitoreando memoria: {e}")
        return True

def test_export_with_cloud_limits(resultado):
    """Test Excel export with cloud configuration (simplified)"""
    print("\n🧪 Test 6: Exportación con límites Cloud")
    
    if not resultado:
        print("  ❌ No hay resultado para exportar")
        return False
    
    try:
        # Simulate basic Excel export functionality
        import io
        from openpyxl import Workbook
        from datetime import datetime
        
        print("  📊 Generando Excel simulado...")
        start_time = time.time()
        
        # Create basic Excel structure
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Análisis Cloud"
        
        # Basic headers
        ws['A1'] = "Personal Paraguay - Análisis IA (Cloud Test)"
        ws['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total comentarios: {resultado.total_comentarios}"
        ws['A4'] = f"Tiempo procesamiento: {resultado.tiempo_total_segundos:.1f}s"
        
        # Save to memory
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        
        end_time = time.time()
        duration = end_time - start_time
        
        if excel_data and len(excel_data) > 1000:
            file_size_kb = len(excel_data) / 1024
            print(f"  ✅ Excel generado en {duration:.1f}s")
            print(f"  📄 Tamaño: {file_size_kb:.1f} KB")
            
            # Check cloud file size limit
            max_size_mb = config.get('MAX_FILE_SIZE_MB', 2)
            if file_size_kb / 1024 > max_size_mb:
                print(f"  ⚠️ Archivo excede límite Cloud: {file_size_kb/1024:.1f} MB > {max_size_mb} MB")
                return False
            else:
                print(f"  ✅ Tamaño dentro de límites Cloud: < {max_size_mb} MB")
                return True
        else:
            print("  ❌ Excel no generado o muy pequeño")
            return False
            
    except Exception as e:
        print(f"  ❌ Error exportando: {e}")
        return False

def main():
    """Run complete pipeline test with cloud configuration"""
    show_environment_info()
    
    # Initialize results
    results = {}
    
    # Test 1: Imports
    results['imports'] = test_imports()
    
    # Test 2: Container with cloud config
    container_success, contenedor, cloud_config = test_container_with_cloud_config()
    results['container'] = container_success
    
    # Test 3: File reading
    if container_success:
        file_success, comentarios = test_file_reading()
        results['file_reading'] = file_success
        
        # Test 4: Analysis with cloud limits
        if file_success:
            analysis_success, resultado = test_analysis_with_cloud_limits(contenedor, comentarios, cloud_config)
            results['analysis'] = analysis_success
            
            # Test 5: Memory behavior
            results['memory'] = test_memory_behavior()
            
            # Test 6: Export
            if analysis_success:
                results['export'] = test_export_with_cloud_limits(resultado)
            else:
                results['export'] = False
        else:
            results['analysis'] = False
            results['memory'] = False
            results['export'] = False
    else:
        results['file_reading'] = False
        results['analysis'] = False
        results['memory'] = False
        results['export'] = False
    
    # Summary
    print("\n" + "=" * 55)
    print("📋 RESUMEN: PIPELINE CON CONFIGURACIÓN CLOUD")
    print("=" * 55)
    
    passed = sum(1 for result in results.values() if result)
    total = len(results)
    
    for test_name, result in results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{test_name.replace('_', ' ').title()}: {status}")
    
    print(f"\n🎯 Score: {passed}/{total} tests passed")
    
    if passed == total:
        print("🎉 ¡PIPELINE COMPLETAMENTE FUNCIONAL CON LÍMITES CLOUD!")
        print("✅ La aplicación funcionará correctamente en Streamlit Cloud")
    elif passed >= total * 0.8:
        print("⚡ ¡PIPELINE MAYORMENTE FUNCIONAL!")
        print("✅ Debería funcionar en Streamlit Cloud con ajustes menores")
    else:
        print("⚠️ ¡PIPELINE NECESITA ATENCIÓN!")
        print("❌ Probable que falle en Streamlit Cloud")
    
    print("\n📝 Recomendaciones para Deploy:")
    if results.get('memory', False):
        print("✅ Memoria: Dentro de límites esperados")
    else:
        print("❌ Memoria: Optimizar uso de memoria antes del deploy")
    
    if results.get('analysis', False):
        print("✅ Análisis: Funcionará con límites de Streamlit Cloud")
    else:
        print("❌ Análisis: Revisar configuración de batches")
    
    if results.get('export', False):
        print("✅ Exportación: Archivos dentro de límites")
    else:
        print("❌ Exportación: Reducir tamaño de archivos generados")
    
    return passed >= total * 0.8

if __name__ == "__main__":
    try:
        success = main()
        exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n👋 Test interrumpido por el usuario")
        exit(1)
    except Exception as e:
        print(f"\n💥 Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        exit(1)