"""
Chunked File Processor for Memory-Efficient Large File Handling
FIX P0 #8: Implement true streaming/chunked reading for large Excel files
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Generator, Optional, Dict, Any
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import gc

logger = logging.getLogger(__name__)


class ChunkedFileProcessor:
    """Process large files in memory-efficient chunks"""
    
    def __init__(self, chunk_size: int = 1000, max_memory_mb: int = 100):
        """
        Initialize chunked processor
        
        Args:
            chunk_size: Number of rows per chunk
            max_memory_mb: Maximum memory usage per chunk in MB
        """
        self.chunk_size = chunk_size
        self.max_memory_mb = max_memory_mb
        
    def read_excel_chunks(self, file_path: Path, sheet_name: Any = 0) -> Generator[pd.DataFrame, None, None]:
        """
        Read Excel file in chunks using openpyxl for true streaming
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet to read (default: first sheet)
            
        Yields:
            DataFrame chunks
        """
        logger.info(f"Starting chunked reading of {file_path.name}")
        
        try:
            # Use openpyxl in read_only mode for memory efficiency
            workbook = openpyxl.load_workbook(
                filename=str(file_path),
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            # Get the specified sheet
            if isinstance(sheet_name, int):
                sheet = workbook.worksheets[sheet_name]
            else:
                sheet = workbook[sheet_name] if sheet_name else workbook.active
                
            # Get headers from first row
            headers = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(cell) if cell is not None else f'Column_{i}' 
                          for i, cell in enumerate(row)]
                break
                
            if not headers:
                raise ValueError("No headers found in Excel file")
                
            # Read data in chunks
            chunk_data = []
            row_count = 0
            total_rows = 0
            
            # Start from row 2 (after headers)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Convert row to list, handling None values
                row_data = [cell for cell in row]
                chunk_data.append(row_data)
                row_count += 1
                total_rows += 1
                
                # Yield chunk when size reached
                if row_count >= self.chunk_size:
                    # Create DataFrame from chunk
                    chunk_df = pd.DataFrame(chunk_data, columns=headers)
                    
                    # Log progress
                    logger.debug(f"Processing chunk: rows {total_rows - row_count + 1}-{total_rows}")
                    
                    yield chunk_df
                    
                    # Clear chunk data and force garbage collection
                    chunk_data = []
                    row_count = 0
                    gc.collect()
                    
            # Yield final chunk if there's remaining data
            if chunk_data:
                chunk_df = pd.DataFrame(chunk_data, columns=headers)
                logger.debug(f"Processing final chunk: rows {total_rows - row_count + 1}-{total_rows}")
                yield chunk_df
                
            # Close workbook
            workbook.close()
            logger.info(f"Completed chunked reading: {total_rows} total rows processed")
            
        except Exception as e:
            logger.error(f"Error in chunked Excel reading: {str(e)}")
            raise
            
    def read_csv_chunks(self, file_path: Path, **kwargs) -> Generator[pd.DataFrame, None, None]:
        """
        Read CSV file in chunks
        
        Args:
            file_path: Path to CSV file
            **kwargs: Additional arguments for pd.read_csv
            
        Yields:
            DataFrame chunks
        """
        logger.info(f"Starting chunked reading of CSV {file_path.name}")
        
        # Set default parameters
        read_kwargs = {
            'chunksize': self.chunk_size,
            'encoding': 'utf-8',
            'on_bad_lines': 'skip'
        }
        read_kwargs.update(kwargs)
        
        try:
            # Use pandas built-in chunking for CSV
            chunk_reader = pd.read_csv(file_path, **read_kwargs)
            
            for i, chunk in enumerate(chunk_reader):
                logger.debug(f"Processing CSV chunk {i + 1}")
                yield chunk
                
                # Force garbage collection periodically
                if (i + 1) % 10 == 0:
                    gc.collect()
                    
        except Exception as e:
            logger.error(f"Error in chunked CSV reading: {str(e)}")
            raise
            
    def process_dataframe_chunks(self, df: pd.DataFrame) -> Generator[pd.DataFrame, None, None]:
        """
        Process an existing DataFrame in chunks
        
        Args:
            df: DataFrame to process
            
        Yields:
            DataFrame chunks
        """
        total_rows = len(df)
        
        for start_idx in range(0, total_rows, self.chunk_size):
            end_idx = min(start_idx + self.chunk_size, total_rows)
            
            # Use iloc for efficient slicing
            chunk = df.iloc[start_idx:end_idx].copy()
            
            logger.debug(f"Processing DataFrame chunk: rows {start_idx + 1}-{end_idx}")
            yield chunk
            
            # Force garbage collection periodically
            if end_idx % (self.chunk_size * 10) == 0:
                gc.collect()
                
    def estimate_memory_usage(self, df: pd.DataFrame) -> float:
        """
        Estimate memory usage of DataFrame in MB
        
        Args:
            df: DataFrame to estimate
            
        Returns:
            Memory usage in MB
        """
        return df.memory_usage(deep=True).sum() / (1024 * 1024)
        
    def optimize_chunk_size(self, sample_df: pd.DataFrame) -> int:
        """
        Optimize chunk size based on sample data and memory limit
        
        Args:
            sample_df: Sample DataFrame for estimation
            
        Returns:
            Optimized chunk size
        """
        if len(sample_df) == 0:
            return self.chunk_size
            
        # Estimate memory per row
        memory_per_row = self.estimate_memory_usage(sample_df) / len(sample_df)
        
        # Calculate optimal chunk size
        if memory_per_row > 0:
            optimal_size = int(self.max_memory_mb / memory_per_row)
            # Ensure reasonable bounds
            return max(100, min(optimal_size, 10000))
        
        return self.chunk_size
        
    def write_excel_chunks(self, chunks: Generator[pd.DataFrame, None, None], 
                          output_path: Path, sheet_name: str = 'Sheet1'):
        """
        Write DataFrame chunks to Excel file efficiently
        
        Args:
            chunks: Generator of DataFrame chunks
            output_path: Output file path
            sheet_name: Name of the sheet to write
        """
        logger.info(f"Starting chunked writing to {output_path.name}")
        
        first_chunk = True
        start_row = 0
        
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            for chunk in chunks:
                if first_chunk:
                    # Write first chunk with headers
                    chunk.to_excel(writer, sheet_name=sheet_name, index=False)
                    first_chunk = False
                    start_row = len(chunk) + 1  # +1 for header row
                else:
                    # Append subsequent chunks without headers
                    chunk.to_excel(writer, sheet_name=sheet_name, 
                                 startrow=start_row, index=False, header=False)
                    start_row += len(chunk)
                    
                logger.debug(f"Written {len(chunk)} rows to Excel")
                
        logger.info(f"Completed chunked writing: {start_row - 1} total rows written")
        
    def write_csv_chunks(self, chunks: Generator[pd.DataFrame, None, None], 
                        output_path: Path, **kwargs):
        """
        Write DataFrame chunks to CSV file efficiently
        
        Args:
            chunks: Generator of DataFrame chunks
            output_path: Output file path
            **kwargs: Additional arguments for to_csv
        """
        logger.info(f"Starting chunked writing to CSV {output_path.name}")
        
        write_kwargs = {
            'index': False,
            'encoding': 'utf-8'
        }
        write_kwargs.update(kwargs)
        
        first_chunk = True
        rows_written = 0
        
        for chunk in chunks:
            if first_chunk:
                # Write first chunk with headers
                chunk.to_csv(output_path, mode='w', **write_kwargs)
                first_chunk = False
            else:
                # Append subsequent chunks without headers
                chunk.to_csv(output_path, mode='a', header=False, **write_kwargs)
                
            rows_written += len(chunk)
            logger.debug(f"Written {len(chunk)} rows to CSV")
            
        logger.info(f"Completed chunked writing: {rows_written} total rows written")